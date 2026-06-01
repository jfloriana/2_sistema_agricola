from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from database import get_db
from sqlalchemy import text
from fastapi.responses import StreamingResponse
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])

@router.get("/cosechas/csv")
def exportar_cosechas_csv(db: Session = Depends(get_db)):
    query = text("SELECT c.fecha_cosecha, c.rendimiento_toneladas, c.calidad FROM cosechas c")
    resultados = db.execute(query).fetchall()
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(["Fecha Cosecha", "Rendimiento (Ton)", "Calidad"])
    
    for row in resultados:
        writer.writerow([row.fecha_cosecha, row.rendimiento_toneladas, row.calidad])
        
    output.seek(0)
    
    # Registro en bitácora
    db.execute(text("INSERT INTO bitacora_sucesos (id_usuario, modulo, accion) VALUES (1, 'Reportes', 'Exportación CSV generada')"))
    db.commit()
    
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="text/csv", 
        headers={"Content-Disposition": "attachment; filename=reporte_cosechas.csv"}
    )

@router.get("/cosechas/excel")
def exportar_cosechas_excel(db: Session = Depends(get_db)):
    query = text("SELECT c.fecha_cosecha, c.rendimiento_toneladas, c.calidad FROM cosechas c ORDER BY c.fecha_cosecha DESC")
    resultados = db.execute(query).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cosechas"

    headers = ["Fecha Cosecha", "Rendimiento (Ton)", "Calidad"]
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row in enumerate(resultados, 2):
        ws.cell(row=row_idx, column=1, value=str(row.fecha_cosecha))
        ws.cell(row=row_idx, column=2, value=float(row.rendimiento_toneladas))
        ws.cell(row=row_idx, column=3, value=row.calidad)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    db.execute(text("INSERT INTO bitacora_sucesos (id_usuario, modulo, accion) VALUES (1, 'Reportes', 'Exportación Excel generada')"))
    db.commit()

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_cosechas.xlsx"}
    )

@router.get("/rendimiento-mensual")
def obtener_rendimiento_mensual(db: Session = Depends(get_db)):
    datos_formateados = {}

    # 1. Pronósticos (se procesan primero, los reales los sobrescriben)
    query_p = text("""
        SELECT 
            p.anio_objetivo,
            p.mes_objetivo,
            cu.nombre_cultivo,
            AVG(p.produccion_estimada) AS total_estimado
        FROM pronosticos p
        JOIN cultivos cu ON p.id_cultivo = cu.id_cultivo
        GROUP BY p.anio_objetivo, p.mes_objetivo, cu.nombre_cultivo
        ORDER BY p.anio_objetivo, p.mes_objetivo
    """)
    for fila in db.execute(query_p).fetchall():
        mes = f"{fila.anio_objetivo}-{fila.mes_objetivo:02d}"
        clave = fila.nombre_cultivo + "_pronostico"
        if mes not in datos_formateados:
            datos_formateados[mes] = {"mes": mes}
        datos_formateados[mes][clave] = round(float(fila.total_estimado), 2)

    # 2. Datos reales (sobrescriben a los pronósticos si existen)
    query_r = text("""
        SELECT 
            TO_CHAR(c.fecha_cosecha, 'YYYY-MM') AS mes,
            cu.nombre_cultivo,
            SUM(c.rendimiento_toneladas) AS total_toneladas
        FROM cosechas c
        JOIN siembras s ON c.id_siembra = s.id_siembra
        JOIN cultivos cu ON s.id_cultivo = cu.id_cultivo
        GROUP BY TO_CHAR(c.fecha_cosecha, 'YYYY-MM'), cu.nombre_cultivo
        ORDER BY mes
    """)
    for fila in db.execute(query_r).fetchall():
        mes = fila.mes
        cultivo = fila.nombre_cultivo
        total = float(fila.total_toneladas)
        if mes not in datos_formateados:
            datos_formateados[mes] = {"mes": mes}
        datos_formateados[mes][cultivo] = round(total, 2)
        # Si había un pronóstico para este cultivo/mes, lo eliminamos
        clave_pronostico = cultivo + "_pronostico"
        datos_formateados[mes].pop(clave_pronostico, None)

    return list(datos_formateados.values())

@router.get("/calidad-cosechas")
def obtener_calidad_cosechas(db: Session = Depends(get_db)):
    query = text("""
        SELECT calidad, SUM(rendimiento_toneladas) as total_toneladas
        FROM cosechas
        GROUP BY calidad
    """)
    resultados = db.execute(query).fetchall()
    
    # Formatear para el PieChart de React
    return [{"name": r.calidad, "value": float(r.total_toneladas)} for r in resultados]

@router.get("/cosechas/json")
def exportar_cosechas_json(db: Session = Depends(get_db)):
    query = text("SELECT c.fecha_cosecha, c.rendimiento_toneladas, c.calidad FROM cosechas c ORDER BY c.fecha_cosecha DESC LIMIT 200")
    resultados = db.execute(query).fetchall()
    
    # Registrar auditoría
    db.execute(text("INSERT INTO bitacora_sucesos (id_usuario, modulo, accion) VALUES (1, 'Reportes', 'Generación de reporte PDF')"))
    db.commit()
    
    return [{"fecha": str(r.fecha_cosecha), "rendimiento": float(r.rendimiento_toneladas), "calidad": r.calidad} for r in resultados]